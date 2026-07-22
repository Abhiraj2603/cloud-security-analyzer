from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font


def generate_excel(scan_data):

    wb = Workbook()

    bold = Font(bold=True)

    # =========================
    # Summary Sheet
    # =========================

    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    summary = scan_data.get("summary", {})

    summary_sheet["A1"] = "Cloud Security Analyzer Report"
    summary_sheet["A1"].font = bold

    summary_sheet["A3"] = "Total Findings"
    summary_sheet["B3"] = summary.get("total_findings", 0)

    summary_sheet["A4"] = "Security Groups"
    summary_sheet["B4"] = summary.get("total_sgs", 0)

    summary_sheet["A5"] = "Risk Score"

    findings = summary.get("total_findings", 0)

    if findings == 0:
        risk = "LOW"
    elif findings < 5:
        risk = "MEDIUM"
    else:
        risk = "HIGH"

    summary_sheet["B5"] = risk

    # =========================
    # Findings Sheet
    # =========================

    findings_sheet = wb.create_sheet("Findings")

    headers = [
        "Severity",
        "Category",
        "Region",
        "Title",
        "Recommendation"
    ]

    for col, header in enumerate(headers, start=1):
        cell = findings_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = bold

    row = 2

    for finding in scan_data.get("findings", []):

        findings_sheet.cell(row=row, column=1).value = finding.get("severity")

        findings_sheet.cell(row=row, column=2).value = finding.get("category")

        findings_sheet.cell(row=row, column=3).value = finding.get("region")

        findings_sheet.cell(row=row, column=4).value = finding.get("title")

        findings_sheet.cell(row=row, column=5).value = finding.get("remediation")

        row += 1

    # Auto size columns
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 5

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output
